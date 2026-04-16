import openpyxl

def modifica_xlsx_per_nome_colonna(file_path, modifiche):
    """
    Modifica valori cercando la colonna per NOME (dalla prima riga).
    
    Args:
        file_path: percorso del file .xlsx (es: 'dati.xlsx')
        modifiche: dict dove:
                   - chiave: nome della colonna (da prima riga)
                   - valore: dict con {numero_riga: nuovo_valore}
    
    Esempio:
        modifiche = {
            'Nome': {2: 'Marco', 5: 'Luca'},
            'Età': {2: 30, 5: 25},
            'Stipendio': {3: 2500, 7: 3000},
        }
        modifica_xlsx_per_nome_colonna('dati.xlsx', modifiche)
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Leggi la prima riga per trovare i nomi delle colonne
    intestazione = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            intestazione[cell.value] = col_idx
    
    # Modifica le celle
    for nome_colonna, righe_valori in modifiche.items():
        if nome_colonna not in intestazione:
            print(f"⚠️ Colonna '{nome_colonna}' non trovata!")
            continue
        
        col_idx = intestazione[nome_colonna]
        
        for riga, valore in righe_valori.items():
            ws.cell(row=riga, column=col_idx).value = valore
            print(f"Modificato {nome_colonna} (riga {riga}): {valore}")
    
    wb.save(file_path)
    print(f"\n✓ File {file_path} aggiornato con successo")


# ESEMPIO DI USO:
if __name__ == "__main__":
    
    # Specifica le modifiche per nome colonna
    modifiche = {
        'FIT_101_linearized': {
            4: 0.12345,
        }
    }
    
    modifica_xlsx_per_nome_colonna('results.xlsx', modifiche)