import torch
import matplotlib.pyplot as plt
import openpyxl
import torch.nn.functional as F
from scipy import signal
import numpy as np
import torchaudio.functional as F_audio
import torch
import torch.nn as nn


def set_params(root, exp_identifier=None, folder_model_101=None, folder_model_101_s2=None, epochs_s1 = 2000, epochs_s2 = 1000, epochs_101_s1 = None, epochs_101_s2 = None, learning_rate = 1e-3, use_noise = False):
    # # # # # # # # Parameters # # # # # # # #
    
    torch.set_default_dtype(torch.float32) 

    #Model
    x0 = torch.tensor([0.01, 0.01])  # Initial state
    input_dim = [1, 1] # input dimensions
    output_dim = [1, 1] # output dimensions


    dim_internal = [8, 8] # [3, 4] # \xi dimension -- number of states of REN
    dim_nl = [8, 8] # [2, 2] # dimension of the square matrix D11 -- number of _non-linear layers_ of the REN

    y_init = torch.tensor([0.0, 0.0])

    IQC_type = ['monotone', 'monotone'] # IQC constraint type: 'l2_gain', 'monotone', 'passive'
    # gamma = torch.tensor([0.3, 0.02])  # for IQC constraints
    gamma = torch.tensor([5, 500])


    ts = 5  # Sampling time (minutes)







    # # # # # # # # Hyperparameters # # # # # # # #
    if learning_rate is None:
        learning_rate = 1e-3
    
    if epochs_s1 is None:
        epochs_s1 = 2000 # 500
    if epochs_s2 is None:
        epochs_s2 = 1000 # 500
    if epochs_101_s1 is None:
        epochs_101_s1 = 2000
    if epochs_101_s2 is None:
        epochs_101_s2 = 10

    # # # # # # # # Data path # # # # # # # #

    redo_save = True
    redo_save_101_I = True
    redo_save_101_M = True

    if exp_identifier is None:
        exp_identifier = '3' # train_batched exp_1
        
    num_days = 30  # 30 2
    
    if folder_model_101 is None or True:
        if exp_identifier <= '10':
            folder_model_101 = f'exp_lin_{exp_identifier}'
        else:
            folder_model_101 = 'baseline' 
    else:
        raise ValueError("folder_model_101 deve essere None o True")

    if folder_model_101_s2 is None or True:
        if exp_identifier  <= '10':
            folder_model_101_s2 = f'exp_lin_{exp_identifier}'
        else:
            folder_model_101_s2 = 'baseline' 
    else:
        raise ValueError("folder_model_101_s2 deve essere None o True")

   

    string_noise = ''
    if use_noise:
        string_noise = '_rwgn'


    data_path = f'{root}/data/train/sc_' + str(num_days) +  'days_identification' + string_noise + '/'
    model_folder = f'{root}/models/SSM/exp_' + exp_identifier + '_' + str(num_days) + 'days' + string_noise + '/'



    return x0, input_dim, output_dim, dim_internal, dim_nl, y_init, IQC_type, gamma, learning_rate, epochs_s1, epochs_s2, epochs_101_s1, epochs_101_s2, data_path, model_folder, redo_save, ts, use_noise, num_days, redo_save_101_I, redo_save_101_M, exp_identifier, folder_model_101, folder_model_101_s2



def ensure_3d(x):
    """ensures that tensors have dimension (batch, time, input_dim)."""
    if x.ndim == 1:
        # Case: sequence 1D -> (1, T, 1)
        x = x.unsqueeze(0).unsqueeze(-1)
    elif x.ndim == 2:
        # Case: batvh or sequence 2D -> (batch, T, 1)
        x = x.unsqueeze(-1)
    return x

def fun_start_controller(train_loader, loaded_parameters, scaler_glucose, scaler_insulin, dataset):

    CGM = dataset.CGM
    sat_e = dataset.sat_e
    
    processed = []
    for batch in train_loader:
        time_batch = batch[-1]

        # Se manca la dimensione batch, aggiungila
        if time_batch.dim() == 1:
            time_batch = time_batch.unsqueeze(0)   # (1, seq_len)

        processed.append(time_batch)

    time_batches = torch.cat(processed, dim=0)

    current_time_index = (time_batches[:, 0].int()).unsqueeze(1)
    previous_starting_index = (time_batches[:,0].int()-1).unsqueeze(1)
    previous_int_duration = previous_starting_index + torch.arange(-loaded_parameters.PID_par.integral_duration, 1) 


    saturation_error_init = scaler_insulin.denormalize(sat_e[previous_starting_index.long()].reshape_as(previous_starting_index))

    y_0 = CGM[current_time_index.long()].reshape_as(current_time_index)
    glucose_PID_init = scaler_glucose.denormalize(CGM[previous_int_duration.long()].reshape_as(previous_int_duration))

    # initial saturation error, string of previous CGM measurament, current CGM measurement
    return saturation_error_init, glucose_PID_init, y_0



def fun_start_controller_simple(train_loader, dataset):

    CGM = dataset.y
    
    processed = []
    for batch in train_loader:
        time_batch = batch[-1]

        # Se manca la dimensione batch, aggiungila
        if time_batch.dim() == 1:
            time_batch = time_batch.unsqueeze(0)   # (1, seq_len)

        processed.append(time_batch)

    time_batches = torch.cat(processed, dim=0)

    current_time_index = (time_batches[:, 0].int()).unsqueeze(1)


    # already normalized
    y_0 = CGM[current_time_index.long()].reshape_as(current_time_index)

    # initial saturation error, string of previous CGM measurament, current CGM measurement
    return y_0

def plot_glucose_insulin(time, insulin=None, meal=None, glucose=None, 
                         predicted_glucose=None, title='Glucose and Insulin vs Time', return_fig = False):
    """
    Plot Glucose/Meal (top) and Insulin (bottom, if present) with dual y-axes.
    Meal points are shown as scatter only when non-zero.
    
    Parameters:
    - time: time array
    - insulin: insulin array (optional)
    - meal: meal array (optional) - plotted as scatter for non-zero values
    - glucose: actual glucose array (optional)
    - predicted_glucose: predicted glucose array (optional)
    - title: plot title
    """
    
    # Determina numero di subplot: 2 solo se ci sono sia insulin che glucose/meal
    has_glucose_or_meal = (glucose is not None or predicted_glucose is not None or meal is not None)
    n_plots = 2 if (insulin is not None and has_glucose_or_meal) else 1
    
    # Usa height_ratios per fare il secondo subplot più stretto
    if n_plots == 2:
        fig, axes = plt.subplots(n_plots, 1, figsize=(10, 3.5 + 1.75), 
                                 gridspec_kw={'height_ratios': [2, 1], 'hspace': 0.05})
    else:
        fig, axes = plt.subplots(n_plots, 1, figsize=(10, 3.5))
    
    if n_plots == 1:
        axes = [axes]
    
    # ===== PRIMO PLOT: Meal/Glucose (o solo Insulin) =====
    ax1 = axes[0]
    
    # Se c'è il secondo subplot, non mettere xticks visibili sul primo
    if n_plots == 2:
        ax1.tick_params(axis='x', labelbottom=False)
    else:
        ax1.set_xlabel('Time step')
    
    # Se solo insulin, mostra solo insulin nel primo plot
    if insulin is not None and n_plots == 1:
        ax1.set_xlabel('Time step')
        ax1.plot(time, insulin, color='tab:red', label='Insulin', zorder=1, linewidth=2)
        ax1.set_ylabel('Insulin (u1)', color='tab:red')
        ax1.tick_params(axis='y', labelcolor='tab:red')
        ax1.spines['left'].set_color('red')
        ax1.grid(True, alpha=0.3)
        ax1.legend(loc='upper right', fontsize=10)
        
        # Ylim dinamico basato su max insulina
        insulin_max = insulin.max()
        ylim_options = [2.5, 5, 7.5, 10, 15, 20]
        ylim = next((y for y in ylim_options if y >= insulin_max), ylim_options[-1])
        ax1.set_ylim(-1, ylim)
        
        ax1.set_title(title, fontsize=14, fontweight='bold')
    else:
        # Meal come scatter (solo punti non-zero)
        if meal is not None:
            # Filtra solo i valori non-zero
            meal_nonzero_mask = meal != 0
            meal_time = time[meal_nonzero_mask]
            meal_values = meal[meal_nonzero_mask]
            
            ax1.scatter(meal_time, meal_values, color='mediumseagreen', label='Meal', 
                       zorder=2, s=80, alpha=0.8, edgecolors='darkgreen', linewidth=1.5)
            ax1.set_ylabel('Meal (mg)', color='mediumseagreen')
            ax1.tick_params(axis='y', labelcolor='mediumseagreen')
            ax1.spines['left'].set_color('mediumseagreen')
        
        # Glucose a destra
        ax2 = ax1.twinx()
        ax2.set_ylabel('Glucose (mg/dL)', color='tab:blue')
        
        if predicted_glucose is not None:
            ax2.plot(time, predicted_glucose, color='darkblue', label='Predicted Glucose', 
                    zorder=20, linewidth=2, alpha=0.7)
        
        if glucose is not None:
            ax2.plot(time, glucose, color='cornflowerblue', label='Glucose', 
                    zorder=10, linewidth=2, alpha=0.7)
        
        ax2.tick_params(axis='y', labelcolor='tab:blue')
        ax2.spines['right'].set_color('blue')
        
        ax1.grid(True, alpha=0.3)
        
        # Legenda
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax2.legend(lines1 + lines2, labels1 + labels2, loc='upper right', fontsize=10)
        
        ax1.set_title(title, fontsize=14, fontweight='bold')
        if meal is not None:
            ax1.set_ylim(-0.1, max([max(meal) * 1.1, 3]))
        # Limiti asse destro: 30-300 ma adatta se i dati sforano
        if glucose is not None or predicted_glucose is not None:
            min_val = 30
            max_val = 300
            
            if glucose is not None and predicted_glucose is not None:
                data_min = min(glucose.min(), predicted_glucose.min())
                data_max = max(glucose.max(), predicted_glucose.max())
            elif glucose is not None:
                data_min = glucose.min()
                data_max = glucose.max()
            else:
                data_min = predicted_glucose.min()
                data_max = predicted_glucose.max()
            
            # Se i dati sforano, adatta i limiti
            if data_min < min_val:
                min_val = data_min * 0.95
            if data_max > max_val:
                max_val = data_max * 1.05
            
            ax2.set_ylim(min_val, max_val)
    
    # ===== SECONDO PLOT: Insulin (se presente insieme a glucose/meal) =====
    if insulin is not None and n_plots == 2:
        ax3 = axes[1]
        ax3.set_xlabel('Time step')
        ax3.plot(time, insulin, color='tab:red', label='Insulin', zorder=1, linewidth=2)
        ax3.set_ylabel('Insulin (u1)', color='tab:red')
        ax3.tick_params(axis='y', labelcolor='tab:red')
        ax3.spines['left'].set_color('red')
        ax3.grid(True, alpha=0.3)
        ax3.legend(loc='upper right', fontsize=10)
        
        # Ylim dinamico basato su max insulina
        insulin_max = insulin.max()
        ylim_options = [1.5, 2.5, 5, 7.5, 10, 15, 20]
        
        # Trova il primo valore >= insulin_max
        ylim = next((y for y in ylim_options if y >= insulin_max), ylim_options[-1])
        
        ax3.set_ylim(-0.5, ylim)
    
    fig.tight_layout()
    
    if return_fig is True:
        return fig
    
    plt.show()

    
    
def modify_xlsx_row_and_column(file_path, modifiche):
    """
    Modifica valori cercando la colonna per NOME (dalla prima riga).
    Crea automaticamente le colonne mancanti nella posizione disponibile.
    
    Args:
        file_path: percorso del file .xlsx (es: 'dati.xlsx')
        modifiche: dict dove:
                   - chiave: nome della colonna (da prima riga)
                   - valore: dict con {numero_riga: nuovo_valore}
    
    Esempio:
        modifiche = {
            'Nome': {2: 'Marco', 5: 'Luca'},
            'Età': {2: 30, 5: 25},
            'Stipendio': {3: 2500, 7: 3000},
        }
        modify_xlsx_row_and_column('dati.xlsx', modifiche)
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Leggi la prima riga per trovare i nomi delle colonne
    intestazione = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            intestazione[cell.value] = col_idx
    
    # Trova la prossima colonna disponibile
    prossima_col_libera = max(intestazione.values()) + 1 if intestazione else 1
    
    # Crea le colonne mancanti nell'ordine in cui le passi
    for nome_colonna in modifiche.keys():
        if nome_colonna not in intestazione:
            ws.cell(row=1, column=prossima_col_libera).value = nome_colonna
            intestazione[nome_colonna] = prossima_col_libera
            print(f"✓ Colonna '{nome_colonna}' creata nella posizione {prossima_col_libera}")
            prossima_col_libera += 1
    
    # Modifica le celle
    for nome_colonna, righe_valori in modifiche.items():
        col_idx = intestazione[nome_colonna]
        
        for riga, valore in righe_valori.items():
            ws.cell(row=riga, column=col_idx).value = valore
            print(f"Modificato {nome_colonna} (riga {riga}): {valore}")
    
    wb.save(file_path)
    print(f"\n✓ File {file_path} aggiornato con successo")
    
def FIT_formula(y_true, y_hat):
    """
    Calcola il FIT basato sulla formula:
    FIT = 100 * (1 - (||y - y_hat|| / ||y - mean(y)||))
    
    Args:
        y_true: array dei valori reali
        y_hat: array dei valori predetti
    
    Returns:
        FIT in percentuale
    """
    import numpy as np

    
    numerator = np.linalg.norm(y_hat - y_true, ord=2)
    denominator = np.linalg.norm(y_true - np.mean(y_true), ord=2)
    FIT = 100 * (1 - numerator / denominator)
    
    return FIT

def moving_average_online(x, window_size, weights_mode='uniform', custom_weights=None):
    """
    Weighted moving average causale (online).
    
    Args:
        x: [batch, time, features]
        window_size: dimensione della finestra
        weights_mode: 'uniform' o 'custom'
        custom_weights: tensore di pesi se weights_mode='custom' [window_size]
    
    Returns:
        [batch, time, features]
    """
    # x: [batch, time, features]
    batch, time, features = x.shape
    device = x.device
    
    if weights_mode == 'uniform':
        weights = torch.ones(window_size, device=device) / window_size
    elif weights_mode == 'custom':
        if custom_weights is None:
            raise ValueError("custom_weights deve essere fornito quando weights_mode='custom'")
        weights = custom_weights.to(device)
        if weights.shape[0] != window_size:
            raise ValueError(f"custom_weights deve avere lunghezza {window_size}")
        weights = weights  
    else:
        raise ValueError(f"weights_mode '{weights_mode}' non riconosciuto")
    
    # Prendi il primo elemento temporale
    first_element = x[:, 0:1, :]  # [batch, 1, features]
    
    # Ripetilo window_size - 1 volte
    padding = first_element.repeat(1, window_size - 1, 1)  # [batch, window_size-1, features]
    
    # Concatena il padding con l'input originale
    x_padded = torch.cat([padding, x], dim=1)  # [batch, time + window_size - 1, features]
    
    # Permuta per convoluzione
    x_padded = x_padded.permute(0, 2, 1)  # [batch, features, time + window_size - 1]
    
    # Crea il kernel di convoluzione
    # weights: [window_size] -> [1, 1, window_size] per conv1d
    kernel = weights.view(1, 1, window_size)
    
    # Replica il kernel per ogni feature
    kernel = kernel.repeat(features, 1, 1)  # [features, 1, window_size]
    
    # Applica convoluzione (groups=features per mantenere le features separate)
    result = F.conv1d(x_padded, kernel, groups=features, stride=1)
    
    return result.permute(0, 2, 1)  # [batch, time, features]



class LowPassFilter(nn.Module):
    def __init__(self, T_sample=5, T_cut=90, order=4, mode='real_poles'):
        super().__init__()
        
        f_sample = 1.0 / (T_sample * 60)
        f_cut = 1.0 / (T_cut * 60)
        f_nyq = 0.5 * f_sample
        Wn = f_cut / f_nyq
        
        self.order = order
        self.mode = mode
        self.batch_size = None

        if mode == 'butter':
            b_np, a_np = signal.butter(order, Wn, btype='low')
        elif mode == 'real_poles':
            p = np.exp(-2 * np.pi * f_cut / f_sample)
            a_poly = np.array([1.0])
            for _ in range(order):
                a_poly = np.convolve(a_poly, np.array([1.0, -p]))
            a_np = a_poly
            b_np = np.zeros(order + 1)
            b_np[0] = (1 - p)**order 
        else:
            raise ValueError("Mode deve essere 'butter' o 'real_poles'")

        self.register_buffer('b', torch.tensor(b_np, dtype=torch.float32))
        self.register_buffer('a_neg', torch.tensor(a_np, dtype=torch.float32))
        
    def _init_buffers(self, batch_size, features, device):
        """Crea i buffer vuoti (verranno riempiti dal reset)"""
        for name in ['x_hist', 'y_hist']:
            if hasattr(self, name) and name not in self._buffers:
                delattr(self, name)  # rimuove attributo "sporco"
            elif name in self._buffers:
                del self._buffers[name]  # rimuove buffer esistente
        self.register_buffer('x_hist', torch.zeros(batch_size, self.order, features, device=device))
        self.register_buffer('y_hist', torch.zeros(batch_size, self.order, features, device=device))
        self.batch_size = batch_size
        
    def reset(self, initial_value):
        """
        Inizializza i buffer con il valore iniziale fornito.
        initial_value: [batch_size, features]
        """
        # Creiamo un tensore [batch_size, order, features] espandendo il valore iniziale
        # initial_value[:, None, :] aggiunge la dimensione 'order'
        fill_value = initial_value[:, None, :].expand(-1, self.order, -1).contiguous()
        
        self.x_hist = fill_value.clone()
        self.y_hist = fill_value.clone()

    def forward(self, x_n):
        # 1. Parte MA
        term_b = self.b[0] * x_n + torch.sum(self.b[None, 1:, None] * self.x_hist, dim=1)
        
        # 2. Parte AR
        term_a = torch.sum(self.a_neg[None, 1:, None] * self.y_hist, dim=1)
        
        y_n = term_b - term_a
        
        # 3. Shift registri
        self.x_hist = torch.cat([x_n[:, None, :], self.x_hist[:, :-1, :]], dim=1)
        self.y_hist = torch.cat([y_n[:, None, :], self.y_hist[:, :-1, :]], dim=1)
        
        return y_n

    def run(self, x_seq):
        """Processa la sequenza chiamando automaticamente il reset sul primo valore"""
        batch_size, time_steps, features = x_seq.shape
        device = x_seq.device
        
        # Inizializzazione struttura se batch_size cambia
        if self.batch_size != batch_size:
            self._init_buffers(batch_size, features, device)
        
        # AUTO-RESET: riempie x_hist e y_hist con x_seq[:, 0, :]
        self.reset(x_seq[:, 0, :])
        
        y_total = torch.zeros_like(x_seq)
        for t in range(time_steps):
            y_total[:, t, :] = self.forward(x_seq[:, t, :])
        
        return y_total
    
    def __call__(self, x_seq):
        return self.run(x_seq)